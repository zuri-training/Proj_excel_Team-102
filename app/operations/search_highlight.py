from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from datetime import datetime

def search_and_highlight(dir, file, search_keyword):
    
    wb = load_workbook(filename=file)
    
    ws = wb.active

    num_match = 0

    for rows in ws.rows:
        for cell in rows:
            if cell.data_type == "n":
                try:
                    check_value = float(search_keyword)
                    if cell.value == check_value: 
                        cell.fill = PatternFill("solid", fgColor="D3E06E")
                        num_match += 1
                except Exception as ex:
                    pass
            else:
                if str(cell.value) == str(search_keyword): 
                    cell.fill = PatternFill("solid", fgColor="D3E06E")
                    num_match += 1

    highlighted_filename = f"{abs(hash(datetime.now()))}_highlighted.xlsx"
    highlighted = f"{dir}/{highlighted_filename}"

    wb.save(highlighted)

    return num_match, highlighted_filename
