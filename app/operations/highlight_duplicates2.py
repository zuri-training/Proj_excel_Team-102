from datetime import datetime
from functools import reduce

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def highlight_duplicates2(dir, file, columns):
    if columns:
        columns = columns.split(",")

    wb = load_workbook(filename=file)

    ws = wb.active

    headers = ws[1]
    headers = [cell.value for cell in headers]

    if columns:
        column_check = {i:{} for i,value in enumerate(headers) if value in columns }
    else:
        column_check = {i:{} for i,value in enumerate(headers)}

    rows_count = reduce(lambda a , b: a + b, [1 for _ in ws.rows])

    rows = ws[2:rows_count]

    duplicate_rows = {}

    for i, row in enumerate(rows, start=2):
        for j, cell in enumerate(row):
            if j in column_check:
                if cell.value in column_check[j]:
                    if i not in duplicate_rows:
                        duplicate_rows[i] = True

                        num_cols = len(headers) + 1
                        last_col = 64 + num_cols

                        for current_col in range(65, last_col):
                            ws[f"{chr(current_col)}{i}"].fill = PatternFill("solid", fgColor="D3E06E")
                    column_check[j][cell.value] += 1
                else:
                    column_check[j][cell.value] = 0

    highlighted_filename = f"{abs(hash(datetime.now()))}_highlighted.xlsx"
    highlighted = f"{dir}/{highlighted_filename}"

    wb.save(highlighted)

    swb = Workbook()

    sws = swb.active

    sws.append(["Unique Columns", "Repeated Values", "Number of Rows"])

    for column_index, data in column_check.items():
        repeated_values = [f"{k}({v})" for k,v in data.items() if v > 0]
        repeated_values = ",".join(repeated_values)

        repeated_rows = sum([v for k, v in data.items() if v > 0])

        sws.append([headers[column_index], repeated_values, repeated_rows])

    analytics_filename = f"{abs(hash(datetime.now()))}_analytics.xlsx"
    analytics = f"{dir}/{analytics_filename}"

    swb.save(analytics)

    return len(duplicate_rows), highlighted_filename, analytics_filename