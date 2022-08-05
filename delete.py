from functools import reduce
from email import header
from operator import index
from openpyxl import Workbook, load_workbook
import warnings
warnings.simplefilter("ignore")

def remove_dups(file, columns):
    if columns:
        columns = columns.split(',')

    wb = load_workbook('stud.xlsx')
    ws = wb.active

    headers = ws[1]
    headers = [cell.value for cell in headers]
    print(cell.value)

    if columns:
        column_check = {i:{} for i,value in enumerate(headers) if value in columns }
    else:
        column_check = {i:{} for i,value in enumerate(headers)}

    rows_count = reduce(lambda a , b: a + b, [1 for _ in ws.rows])
    rows = ws[2:rows_count]
    duplicate_rows = {}

    for i, row in enumerate(rows, start=2):
        for j, cell in enumerate(row):
            if j in column_check:
                if cell.value in column_check[j]:
                    duplicate_rows[i] = duplicate_rows.get(i, [])
                    duplicate_rows[i].append(headers[j])
                else:
                    column_check[j][cell.value] = True

    new_data = [headers]
    new_data += [[cell.value for cell in row] for i, row in enumerate(rows, start=2) if i not in duplicate_rows]
    ws.delete_rows(1, rows_count)

    for data in new_data:
        ws.append(data)
    without_duplicates = f"{file}without_duplicates2.xlsx"
    wb.save(without_duplicates)
    return without_duplicates


remove_dups('stud2.xlsx', 'Student ID,Name,Course')



