from datetime import datetime
from os import path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def diff_checker(dir, file1, file2):
    if not path.exists(file1):
        print(f"Could not resolve {file1}"); return        
    if not path.exists(file2):
        print(f"Could not resolve {file2}"); return

    wb1, wb2 = load_workbook(filename=file1), load_workbook(filename=file2)

    ws1, ws2 = wb1.active, wb2.active

    num_mismatch = 0

    for row in ws1.rows:
        for cell in row:
            if cell.value != ws2[cell.coordinate].value:
                num_mismatch += 1
                cell.fill = PatternFill("solid", fgColor="D3E06E")
                ws2[cell.coordinate].fill = PatternFill("solid", fgColor="D3E06E")

    highlighted_file1name, highlighted_file2name = f"{abs(hash(datetime.now()))}_highlighted_file1.xlsx", f"{abs(hash(datetime.now()))}_highlighted_file2.xlsx"

    wb1.save(f"{dir}{highlighted_file1name}"); wb2.save(f"{dir}{highlighted_file2name}")

    return num_mismatch, highlighted_file1name, highlighted_file2name