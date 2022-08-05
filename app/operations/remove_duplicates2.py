from datetime import datetime
from functools import reduce

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def remove_duplicates2(dir, file, columns):
    if columns:
        columns = columns.split(",")

    wb = load_workbook(filename=file)

    ws = wb.active

    headers = ws[1]
    headers = [cell.value for cell in headers]

    if columns:
        column_check = {i:{} for i,value in enumerate(headers) if value in columns }
    else:
        column_check = {i:{} for i,value in enumerate(headers)}

    rows_count = reduce(lambda a , b: a + b, [1 for _ in ws.rows])

    rows = ws[2:rows_count]

    duplicate_rows = {}

    for i, row in enumerate(rows, start=2):
        for j, cell in enumerate(row):
            if j in column_check:
                if cell.value in column_check[j]:
                    duplicate_rows[i] = duplicate_rows.get(i, [])
                    duplicate_rows[i].append(headers[j])
                else:
                    column_check[j][cell.value] = True
        

    dwb = Workbook()

    dws = dwb.active

    dws.append(headers + ["Duplicated Columns"])

    num_cols = len(headers) + 1
    last_col = chr(64 + num_cols)

    dws[f"{last_col}1"].fill = PatternFill("solid", fgColor="D3E06E")

    current_row = 1

    for row_index, duplicate_cols in duplicate_rows.items():
        data = [cell.value for cell in ws[row_index]]
        data.append(",".join(duplicate_cols))
        dws.append(data)

        current_row += 1
        dws[f"{last_col}{current_row}"].fill = PatternFill("solid", fgColor="D3E06E");

    duplicates_filename = f"{abs(hash(datetime.now()))}_duplicates.xlsx"
    duplicates = f"{dir}/{duplicates_filename}"

    dwb.save(duplicates)

    new_data = [headers]
    new_data += [[cell.value for cell in row] for i, row in enumerate(rows, start=2) if i not in duplicate_rows]

    ws.delete_rows(1, rows_count)

    for data in new_data:
        ws.append(data)

    without_duplicates_filename = f"{abs(hash(datetime.now()))}_without_duplicates.xlsx"
    without_duplicates = f"{dir}/{without_duplicates_filename}"

    wb.save(without_duplicates)

    return len(duplicate_rows), duplicates_filename, without_duplicates_filename