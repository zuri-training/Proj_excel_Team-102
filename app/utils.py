import os
from passlib.context import CryptContext
from math import floor
from datetime import datetime

import aiofiles

from openpyxl import load_workbook, Workbook

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def get_password_hash(password):
    return pwd_context.hash(password)

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_pagination(num_records, division):
    if not num_records:
        return 0
    pages = floor(num_records/division)
    return pages + 1 if (num_records % division) > 0 else pages

def get_skip(page, division):
    return (page - 1) * division

def get_unique_dir():
    unique_dirname = abs(hash(datetime.now()))
    unique_dirname = f"static/files/{unique_dirname}"

    return unique_dirname

def get_unique_filepath(filename, dirname):
    unique_prefix = abs(hash(datetime.now()))
    file_path = f"{dirname}/{unique_prefix}_{filename}"

    return file_path

async def upload_file(filepath, file):
    async with aiofiles.open(filepath, 'wb') as out_file:
        while content := await file.read(1024):  
            await out_file.write(content)

def create_excel_from_csv(filepath):

    wb = Workbook()

    ws = wb.active

    with open(filepath, "r") as file:
        lines = file.readlines()

        for line in lines:
            ws.append(line.split(","))

    wb.save(f"{filepath[:-4]}.xlsx")

    os.remove(filepath)

def get_excel_contents(filepath):
    workbook = load_workbook(filename=filepath)

    worksheet = workbook.active

    data = []

    for row in worksheet.rows:
        row_data = []
        for cell in row:
            row_data.append({
                "value": cell.value,
                "background": cell.fill.fgColor.rgb
            })
        data.append(row_data)

    return data