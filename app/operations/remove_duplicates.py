from functools import reduce
from openpyxl import load_workbook

from datetime import datetime

def remove_duplicates(dir, file, columns):
    if columns:
        columns = columns.split(',')

    wb = load_workbook(filename=file)

    ws = wb.active

    headers = ws[1]
    headers = [cell.value for cell in headers]

    if columns:
        column_check = {i:{} for i,value in enumerate(headers) if value in columns }
    else:
        column_check = {i:{} for i,value in enumerate(headers)}

    rows_count = reduce(lambda a , b: a + b, [1 for _ in ws.rows])

    rows = ws[2:rows_count]
    
    duplicate_rows = {}

    for i, row in enumerate(rows, start=2):
        for j, cell in enumerate(row):
            if j in column_check:
                if cell.value in column_check[j]:
                    duplicate_rows[i] = duplicate_rows.get(i, [])
                    duplicate_rows[i].append(headers[j])
                else:
                    column_check[j][cell.value] = True

    new_data = [headers]

    new_data += [[cell.value for cell in row] for i, row in enumerate(rows, start=2) if i not in duplicate_rows]
    
    ws.delete_rows(1, rows_count)

    for data in new_data:
        ws.append(data)

    without_duplicates_filename = f"{abs(hash(datetime.now()))}_without_duplicates.xlsx"
    without_duplicates = f"{dir}/{without_duplicates_filename}"

    wb.save(without_duplicates)

    return len(duplicate_rows), without_duplicates_filename




